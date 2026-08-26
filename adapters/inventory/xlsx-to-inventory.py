#!/usr/bin/env python3
"""Convert the AM Motors vehicle sheet into normalized inventory JSON.

Reads the "Машины" worksheet of AM-Motors-справочник-машин.xlsx and writes rows
matching data-schemas/vehicle-inventory-record.schema.json, which is what
adapters/inventory/vehicle-facts-adapter.mjs consumes.

This step only renames and type-normalizes. It never invents a value: a blank
cell stays absent so the adapter emits no fact for it.

    pip install openpyxl jsonschema
    python3 adapters/inventory/xlsx-to-inventory.py \
        --xlsx AM-Motors-справочник-машин.xlsx \
        --out data/inventory/am-motors-vehicles.json
"""
from __future__ import annotations

import argparse
import datetime as dt
import json
from pathlib import Path
import re
import sys

SHEET = "Машины"

# Sheet header -> schema field. Headers not listed here are ignored.
COLUMN_MAP = {
    "ID": "vehicle_id",
    "Марка": "make",
    "Модель": "model",
    "Год": "year",
    "VIN": "vin",
    "Цвет": "color",
    "Пробег (км)": "mileage_km",
    "Двигатель": "engine",
    "Топливо": "fuel",
    "Коробка": "transmission",
    "Привод": "drivetrain",
    "Спецификация": "trim",
    "Состояние": "condition",
    "Аварии/крашеные детали": "accident_history",
    "Сервисная история": "service_history",
    "Владельцев": "owners_count",
    "Цена AED": "price_aed",
    "Мин. цена AED": "min_price_aed",
    "Статус": "status",
    "Мулькия до": "registration_valid_until",
    "Банковский залог": "bank_lien",
    "Ссылка на фото/видео": "media_url",
    "Ссылка на пост": "listing_url",
    "Заметки": "notes",
    "Дата обновления": "updated_at",
}

INTEGER_FIELDS = {"year", "mileage_km", "owners_count"}
NUMBER_FIELDS = {"price_aed", "min_price_aed"}
DATE_FIELDS = {"updated_at", "registration_valid_until"}


def blank(value) -> bool:
    if value is None:
        return True
    if isinstance(value, str):
        return value.strip() in {"", "-", "—"}
    return False


def to_number(value):
    """'85 000 AED' -> 85000. Returns None when nothing numeric is present."""
    if isinstance(value, (int, float)) and not isinstance(value, bool):
        return value
    digits = re.sub(r"[^0-9.,-]", "", str(value)).replace(" ", "")
    if digits.count(",") == 1 and digits.count(".") == 0 and len(digits.split(",")[-1]) != 3:
        digits = digits.replace(",", ".")
    else:
        digits = digits.replace(",", "")
    if not re.fullmatch(r"-?\d+(\.\d+)?", digits or ""):
        return None
    return float(digits)


def to_date(value):
    if isinstance(value, dt.datetime):
        return value.date().isoformat()
    if isinstance(value, dt.date):
        return value.isoformat()
    raw = str(value).strip()
    if re.fullmatch(r"\d{4}-\d{2}-\d{2}", raw):
        return raw
    m = re.fullmatch(r"(\d{2})\.(\d{2})\.(\d{4})", raw)
    if m:
        return f"{m.group(3)}-{m.group(2)}-{m.group(1)}"
    return raw


def convert(xlsx: Path) -> list[dict]:
    from openpyxl import load_workbook

    workbook = load_workbook(xlsx, data_only=True)
    if SHEET not in workbook.sheetnames:
        raise SystemExit(f"worksheet {SHEET!r} not found; sheets are {workbook.sheetnames}")
    sheet = workbook[SHEET]

    rows = sheet.iter_rows(values_only=True)
    headers = [None if h is None else str(h).strip() for h in next(rows)]
    unknown = [h for h in headers if h and h not in COLUMN_MAP]
    if unknown:
        print(f"note: ignoring unmapped columns {unknown}", file=sys.stderr)

    records = []
    for raw_row in rows:
        record = {}
        for header, cell in zip(headers, raw_row):
            field = COLUMN_MAP.get(header or "")
            if field is None or blank(cell):
                continue  # blank stays absent; the adapter emits no fact for it
            if field in INTEGER_FIELDS:
                number = to_number(cell)
                if number is None:
                    continue
                record[field] = int(number)
            elif field in NUMBER_FIELDS:
                number = to_number(cell)
                if number is None:
                    continue
                record[field] = number
            elif field in DATE_FIELDS:
                record[field] = to_date(cell)
            else:
                record[field] = str(cell).strip()
        if record:
            records.append(record)
    return records


def main() -> int:
    parser = argparse.ArgumentParser(description=__doc__.splitlines()[0])
    parser.add_argument("--xlsx", required=True, type=Path)
    parser.add_argument("--out", required=True, type=Path)
    parser.add_argument("--schema", type=Path,
                        default=Path("data-schemas/vehicle-inventory-record.schema.json"))
    args = parser.parse_args()

    records = convert(args.xlsx)

    try:
        import jsonschema
    except ImportError:
        print("note: jsonschema not installed, skipping validation", file=sys.stderr)
    else:
        schema = json.loads(args.schema.read_text(encoding="utf-8"))
        validator = jsonschema.Draft202012Validator(schema)
        failures = 0
        for record in records:
            for error in sorted(validator.iter_errors(record), key=lambda e: list(e.absolute_path)):
                failures += 1
                where = "/".join(str(p) for p in error.absolute_path) or "$"
                print(f"{record.get('vehicle_id', '?')}: {where}: {error.message}", file=sys.stderr)
        if failures:
            raise SystemExit(f"{failures} inventory row(s) do not match the schema")

    args.out.parent.mkdir(parents=True, exist_ok=True)
    args.out.write_text(json.dumps(records, indent=2, ensure_ascii=False) + "\n", encoding="utf-8")
    print(json.dumps({"rows": len(records), "out": str(args.out)}, ensure_ascii=False))
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
